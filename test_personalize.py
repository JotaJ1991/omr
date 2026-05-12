"""
test_personalize.py — Genera PDFs personalizados de prueba

Toma el archivo roster_IETAGRO_ejemplo.xlsx, genera un QR por
estudiante/sesión, rellena el template LaTeX y compila a PDF.

Salida: carpeta `out_test/` con los 6 PDFs (3 estudiantes × 2 sesiones).
"""

import os
import sys
import io
import shutil
import subprocess
from pathlib import Path

# Forzar stdout UTF-8 en Windows
if hasattr(sys.stdout, 'buffer'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import qrcode
from openpyxl import load_workbook

# ── Configuración ──
SIM_ID = 'Mayo_2026_IETAGRO'
ROOT = Path(__file__).parent
TEMPLATES = ROOT / 'templates_latex'
ROSTER = ROOT / 'roster_IETAGRO_ejemplo.xlsx'
OUT = ROOT / 'out_test'
TMP = ROOT / 'out_test' / '_tmp'


def make_qr(data: str, dest: Path):
    """Genera un QR PNG con corrección de errores ALTA (25%) para tolerar
    fotocopias, arrugas y manchas."""
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=10,
        border=2,
    )
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')
    img.save(dest)


def latex_escape(s: str) -> str:
    """Escapa caracteres especiales de LaTeX en strings de usuario."""
    s = str(s or '')
    repl = {
        '\\': r'\textbackslash{}',
        '&':  r'\&',
        '%':  r'\%',
        '$':  r'\$',
        '#':  r'\#',
        '_':  r'\_',
        '{':  r'\{',
        '}':  r'\}',
        '~':  r'\textasciitilde{}',
        '^':  r'\textasciicircum{}',
    }
    for k, v in repl.items():
        s = s.replace(k, v)
    return s


def render(template_path: Path, dest_tex: Path, vars: dict):
    """Sustituye <<VAR>> en el template y escribe en dest_tex."""
    tex = template_path.read_text(encoding='utf-8')
    for k, v in vars.items():
        tex = tex.replace(f'<<{k}>>', str(v))
    dest_tex.write_text(tex, encoding='utf-8')


def compile_latex(tex_path: Path, out_dir: Path) -> Path:
    """Compila el .tex y devuelve la ruta del PDF resultante."""
    out_dir.mkdir(parents=True, exist_ok=True)
    cmd = [
        'pdflatex',
        '-interaction=nonstopmode',
        '-halt-on-error',
        '-output-directory', str(out_dir),
        str(tex_path),
    ]
    r = subprocess.run(cmd, capture_output=True, text=True, cwd=str(tex_path.parent))
    pdf_path = out_dir / (tex_path.stem + '.pdf')
    if r.returncode != 0 or not pdf_path.exists():
        print('--- LaTeX log (últimas 30 líneas) ---')
        print('\n'.join(r.stdout.splitlines()[-30:]))
        raise RuntimeError(f'pdflatex falló para {tex_path.name}')
    return pdf_path


def safe_filename(s: str) -> str:
    s = (s or '').upper().strip()
    keep = ''.join(c if c.isalnum() else '_' for c in s)
    while '__' in keep: keep = keep.replace('__', '_')
    return keep.strip('_')


def main():
    if not ROSTER.exists():
        print(f'ERROR: no encuentro {ROSTER}'); sys.exit(1)
    if OUT.exists(): shutil.rmtree(OUT)
    OUT.mkdir(); TMP.mkdir()

    # Leer roster
    wb = load_workbook(ROSTER, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h or '').strip().lower() for h in rows[0]]
    idx_curso = header.index('curso')
    idx_id    = header.index('identificacion')
    idx_name  = header.index('nombre')
    students = []
    for r in rows[1:]:
        if not r: continue
        curso = str(r[idx_curso] or '').strip()
        ident = str(r[idx_id]    or '').strip()
        name  = str(r[idx_name]  or '').strip()
        if curso and ident and name:
            students.append({'curso': curso, 'id': ident, 'nombre': name})
    print(f'Roster: {len(students)} estudiantes\n')

    sessions = [
        ('1S', TEMPLATES / 'sipagre_1s_personalizado.tex'),
        ('2S', TEMPLATES / 'sipagre_2s_personalizado.tex'),
    ]

    generated = []
    for stu in students:
        for ses, tpl in sessions:
            # 1) Generar QR
            payload = f'SIPAGRE|{SIM_ID}|{ses}|{stu["id"]}|{stu["curso"]}'
            qr_path = TMP / f'qr_{stu["id"]}_{ses}.png'
            make_qr(payload, qr_path)

            # 2) Rellenar template
            short_name = ' '.join(stu['nombre'].split()[:3])  # 3 primeras palabras
            vars = {
                'NOMBRE':   latex_escape(short_name),
                'ID':       latex_escape(stu['id']),
                'CURSO':    latex_escape(stu['curso']),
                'SESION':   ses,
                'QR_PATH':  qr_path.as_posix(),  # LaTeX prefiere forward slashes
            }
            base = f'{stu["id"]}_{safe_filename(stu["nombre"][:20])}_{ses}'
            tex_out = TMP / f'{base}.tex'
            render(tpl, tex_out, vars)

            # 3) Compilar
            print(f'  → {base} ... ', end='', flush=True)
            try:
                pdf = compile_latex(tex_out, TMP)
                # Mover el PDF a out/<curso>/
                curso_dir = OUT / stu['curso']
                curso_dir.mkdir(exist_ok=True)
                dest_pdf = curso_dir / f'{base}.pdf'
                shutil.move(str(pdf), str(dest_pdf))
                generated.append(dest_pdf)
                print('OK')
            except Exception as e:
                print(f'ERROR: {e}')

    # Limpiar archivos auxiliares
    for ext in ('.aux', '.log', '.out', '.tex'):
        for f in TMP.glob(f'*{ext}'):
            try: f.unlink()
            except: pass

    print(f'\n✅ {len(generated)} PDFs generados en {OUT}')
    print('Estructura:')
    for curso_dir in sorted(OUT.iterdir()):
        if curso_dir.is_dir() and curso_dir.name != '_tmp':
            print(f'  {curso_dir.name}/')
            for f in sorted(curso_dir.iterdir()):
                print(f'    {f.name}')


if __name__ == '__main__':
    main()
