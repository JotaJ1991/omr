"""
pdf_personalizer.py — Generación de hojas OMR personalizadas con QR
====================================================================
Lee un roster de estudiantes (XLSX), genera un PDF por estudiante y sesión
con el LaTeX template, embebido con QR único.

Acepta cabeceras de columnas con varios alias comunes:
  - Curso        | "Curso", "CURSO"
  - Documento    | "Documento", "Identificacion", "ID", "Cédula", "Cedula", "DNI"
  - Nombre       | "Nombre", "Nombres", "Apellidos y Nombres"

Salida: ZIP organizado por curso.
"""

import io
import os
import re
import sys
import shutil
import zipfile
import tempfile
import subprocess
from pathlib import Path
from typing import Optional
from concurrent.futures import ProcessPoolExecutor, as_completed

# Importaciones que pueden faltar en algunos despliegues — se manejan con
# mensajes de error claros si fallan.
try:
    import qrcode
    _QRCODE_OK = True
except ImportError:
    _QRCODE_OK = False
    qrcode = None

try:
    from openpyxl import load_workbook
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False
    load_workbook = None

try:
    from pypdf import PdfWriter, PdfReader
    _PYPDF_OK = True
except ImportError:
    _PYPDF_OK = False
    PdfWriter = PdfReader = None


def _check_deps() -> Optional[str]:
    """Devuelve None si todo OK, o un mensaje de error con instrucciones."""
    missing = []
    if not _QRCODE_OK:    missing.append('qrcode')
    if not _OPENPYXL_OK:  missing.append('openpyxl')
    if not _PYPDF_OK:     missing.append('pypdf')
    if missing:
        return (f"Faltan dependencias Python: {', '.join(missing)}. "
                f"En el servidor: pip install {' '.join(missing)}")
    if not shutil.which('pdflatex'):
        return ("pdflatex no está instalado en el servidor. Esta función "
                "requiere LaTeX (MiKTeX o TeXLive) para generar los PDFs. "
                "Si la app corre en Render u otro PaaS sin LaTeX, "
                "ejecuta la generación en tu equipo local donde tengas "
                "MiKTeX instalado.")
    return None


ROOT = Path(__file__).parent
TEMPLATES = ROOT / 'templates_latex'
TPL_1S = TEMPLATES / 'ietagro_1s.tex'
TPL_2S = TEMPLATES / 'ietagro_2s.tex'

# Aliases case-insensitive para cabeceras de columnas
_HEADER_ALIASES = {
    'curso':       ['curso'],
    'documento':   ['documento', 'identificacion', 'identificación', 'id',
                    'cedula', 'cédula', 'dni', 'doc'],
    'nombre':      ['nombre', 'nombres', 'apellidos y nombres',
                    'nombres y apellidos', 'estudiante'],
}


def _norm(s) -> str:
    return str(s or '').strip().lower()


def _map_headers(header_row) -> dict:
    """
    Mapea posiciones de columnas del header del XLSX a campos lógicos.
    Devuelve {'curso': idx, 'documento': idx, 'nombre': idx} o None si falta alguno.
    """
    mapping = {}
    for i, h in enumerate(header_row):
        h_norm = _norm(h)
        for field, aliases in _HEADER_ALIASES.items():
            if h_norm in aliases and field not in mapping:
                mapping[field] = i
                break
    return mapping


def parse_roster_xlsx(xlsx_bytes: bytes) -> dict:
    """
    Lee el XLSX (todas las hojas) y devuelve:
      {'success': True, 'students': [{'curso','id','nombre'}, ...]}
    o {'success': False, 'error': '...'}
    """
    if not _OPENPYXL_OK:
        return {'success': False,
                'error': 'openpyxl no está instalado. Agrega openpyxl al '
                         'requirements.txt y vuelve a desplegar.'}
    try:
        wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    except Exception as e:
        return {'success': False, 'error': f'No se pudo abrir el XLSX: {e}'}

    all_students = []
    sheets_processed = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if not rows: continue
        header = rows[0]
        mapping = _map_headers(header)
        if not all(k in mapping for k in ('curso', 'documento', 'nombre')):
            # Esta hoja no tiene las columnas requeridas
            continue
        sheets_processed.append(sn)
        for r in rows[1:]:
            if not r: continue
            curso = str(r[mapping['curso']] or '').strip()
            ident = str(r[mapping['documento']] or '').strip()
            nombre = str(r[mapping['nombre']] or '').strip()
            if not (curso and ident and nombre): continue
            # Limpiar ID (quitar puntos, comas, espacios)
            ident_clean = re.sub(r'[\s.,\-]', '', ident)
            # Normalizar nombre a mayúsculas
            nombre_up = nombre.upper()
            all_students.append({
                'curso':  curso,
                'id':     ident_clean,
                'nombre': nombre_up,
            })

    if not all_students:
        return {'success': False,
                'error': 'No se encontraron estudiantes válidos. Verifica las '
                         'columnas (Curso, Documento, Nombre).'}
    # Eliminar duplicados por ID
    seen = set()
    unique = []
    for s in all_students:
        if s['id'] in seen: continue
        seen.add(s['id'])
        unique.append(s)
    # Contar por curso
    by_curso = {}
    for s in unique:
        by_curso[s['curso']] = by_curso.get(s['curso'], 0) + 1
    return {
        'success':  True,
        'students': unique,
        'count':    len(unique),
        'by_curso': by_curso,
        'sheets':   sheets_processed,
    }


def _make_qr_png(data: str, dest: Path):
    """Genera un QR PNG con corrección de errores alta (25% tolerancia)."""
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=10, border=2,
    )
    qr.add_data(data); qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')
    img.save(dest)


def _latex_escape(s: str) -> str:
    """Escapa caracteres LaTeX especiales."""
    s = str(s or '')
    repl = {
        '\\': r'\textbackslash{}', '&': r'\&', '%': r'\%', '$': r'\$',
        '#': r'\#', '_': r'\_', '{': r'\{', '}': r'\}',
        '~': r'\textasciitilde{}', '^': r'\textasciicircum{}',
    }
    for k, v in repl.items():
        s = s.replace(k, v)
    return s


def _abbreviate_name(full_name: str, max_chars: int = 28) -> str:
    """Si el nombre excede max_chars, abrevia el segundo nombre a inicial."""
    name = (full_name or '').strip()
    if len(name) <= max_chars: return name
    parts = name.split()
    if len(parts) >= 4:
        # 1er nombre + inicial 2do + apellidos
        abbr = f'{parts[0]} {parts[1][0]}. {" ".join(parts[2:])}'
        if len(abbr) <= max_chars: return abbr
    # Trunc duro
    return name[:max_chars - 1] + '.'


def _safe_filename(s: str) -> str:
    s = (s or '').upper().strip()
    keep = ''.join(c if c.isalnum() else '_' for c in s)
    while '__' in keep: keep = keep.replace('__', '_')
    return keep.strip('_')


def _render_template(tpl_path: Path, dest_tex: Path, vars: dict):
    """Sustituye <<VAR>> en el template y escribe en dest_tex."""
    tex = tpl_path.read_text(encoding='utf-8')
    for k, v in vars.items():
        tex = tex.replace(f'<<{k}>>', str(v))
    dest_tex.write_text(tex, encoding='utf-8')


def _compile_latex(tex_path: Path, out_dir: Path) -> Optional[Path]:
    """Compila el .tex y devuelve la ruta del PDF, o None si falla.

    IMPORTANTE: corre pdflatex DOS VECES. La plantilla usa
    `remember picture` con `current page` para posicionar los marcadores
    fiduciales en las esquinas, lo cual requiere dos pasadas: la primera
    registra las coordenadas en el .aux, la segunda las lee. Sin esto,
    los fiduciales no aparecen y el OMR no puede corregir la perspectiva.

    Usamos nombres RELATIVOS (solo filename) en vez de rutas absolutas
    porque en Windows si la ruta contiene un componente con '~'
    (shortname 8.3, ej. 'ESTUDI~1'), LaTeX lo interpreta como espacio
    no rompible y falla. Al correr con cwd=tmpdir y pasar solo nombres
    de archivo, evitamos ese problema.
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    # Si out_dir == tex_path.parent, no necesitamos --output-directory
    if out_dir.resolve() == tex_path.parent.resolve():
        cmd = ['pdflatex', '-interaction=nonstopmode', '-halt-on-error',
               tex_path.name]
    else:
        cmd = ['pdflatex', '-interaction=nonstopmode', '-halt-on-error',
               '-output-directory', out_dir.name, tex_path.name]
    # Dos pasadas para que `remember picture` registre y use las
    # coordenadas de los marcadores fiduciales.
    for _ in range(2):
        r = subprocess.run(cmd, capture_output=True, text=True,
                           cwd=str(tex_path.parent), timeout=60)
        if r.returncode != 0:
            return None
    pdf = out_dir / (tex_path.stem + '.pdf')
    return pdf if pdf.exists() else None


# Función worker para multiprocessing (debe estar a nivel de módulo para serializar)
def _build_one_pdf(args):
    """Worker: prepara QR + .tex + compila un PDF. Devuelve dict con resultado."""
    (stu, ses, tpl_path, sim_id_safe, tmpdir_str) = args
    tmpdir = Path(tmpdir_str)
    try:
        # Crear sub-dir para este job (evita colisiones de auxiliares aux/log)
        job_dir = tmpdir / f'job_{stu["id"]}_{ses}'
        job_dir.mkdir(parents=True, exist_ok=True)

        qr_payload = f'OMR|{sim_id_safe}|{ses}|{stu["id"]}|{stu["curso"]}'
        qr_path = job_dir / 'qr.png'
        _make_qr_png(qr_payload, qr_path)

        base = f'{stu["id"]}_{_safe_filename(stu["nombre"][:24])}_{ses}'
        tex_out = job_dir / f'{base}.tex'
        vars = {
            'NOMBRE':  _latex_escape(_abbreviate_name(stu['nombre'])),
            'ID':      _latex_escape(stu['id']),
            'CURSO':   _latex_escape(stu['curso']),
            'SESION':  ses,
            'QR_PATH': 'qr.png',  # relativo al job_dir
        }
        _render_template(Path(tpl_path), tex_out, vars)
        pdf = _compile_latex(tex_out, job_dir)
        if not pdf:
            return {'ok': False, 'stu_id': stu['id'], 'ses': ses,
                    'err': 'pdflatex error'}
        return {'ok': True, 'curso': stu['curso'],
                'filename': f'{base}.pdf', 'path': str(pdf)}
    except Exception as e:
        return {'ok': False, 'stu_id': stu['id'], 'ses': ses, 'err': str(e)}


def generate_pdfs_zip(students: list, simulacro_id: str,
                      progress_cb=None, max_workers: int = 4) -> dict:
    """
    Genera los PDFs personalizados (1S + 2S por estudiante) y los MERGEA
    en UN ÚNICO PDF con todas las páginas, ordenadas por curso/estudiante.

    Orden de páginas:
        Curso 11A:
          Estudiante 1: pág 1 = 1S, pág 2 = 2S
          Estudiante 2: pág 3 = 1S, pág 4 = 2S
          ...
        Curso 11B: ...

    Devuelve {'success': True, 'pdf_bytes': bytes, 'count': N_páginas, ...}.
    Mantiene la clave 'zip_bytes' = pdf_bytes por compatibilidad con el
    endpoint pero el contenido ya es un PDF (no un ZIP).

    Usa multiprocessing para paralelizar pdflatex (~4x speedup con 4 workers).
    progress_cb(current, total): callback opcional para progreso.
    """
    dep_err = _check_deps()
    if dep_err:
        return {'success': False, 'error': dep_err}
    if not TPL_1S.exists() or not TPL_2S.exists():
        return {'success': False, 'error': 'Plantillas LaTeX no encontradas.'}

    sim_id_safe = _safe_filename(simulacro_id)
    total = len(students) * 2
    failures = []

    # ── Persistir el roster en Sheets para que el OMR pueda mapear ID→nombre
    #    al detectar el QR. No bloquea el flujo si falla.
    roster_saved = 0
    try:
        from sheets_connector import save_roster_for_simulacro
        rr = save_roster_for_simulacro(sim_id_safe, students)
        roster_saved = rr.get('count', 0) if rr.get('success') else 0
    except Exception:
        roster_saved = 0

    with tempfile.TemporaryDirectory(prefix='ietagro_pdfs_') as tmpdir:
        tmpdir = Path(tmpdir)

        # Construir lista de jobs (uno por estudiante × sesión)
        jobs = []
        for stu in students:
            jobs.append((stu, '1S', str(TPL_1S), sim_id_safe, str(tmpdir)))
            jobs.append((stu, '2S', str(TPL_2S), sim_id_safe, str(tmpdir)))

        out_pdfs = []  # tuples (stu_id, ses, curso, pdf_path)
        done = 0
        with ProcessPoolExecutor(max_workers=max_workers) as ex:
            futures = {ex.submit(_build_one_pdf, j): j for j in jobs}
            for fut in as_completed(futures):
                done += 1
                if progress_cb:
                    progress_cb(done, total)
                try:
                    r = fut.result()
                    if r.get('ok'):
                        # Recuperar info del job para ordenar
                        job = futures[fut]
                        stu_obj = job[0]
                        ses     = job[1]
                        out_pdfs.append((r['curso'], stu_obj.get('nombre', ''),
                                         stu_obj['id'], ses, r['path']))
                    else:
                        failures.append((r.get('stu_id'), r.get('ses'), r.get('err')))
                except Exception as e:
                    failures.append(('?', '?', str(e)))

        if not out_pdfs:
            return {'success': False,
                    'error': 'No se pudo generar ningún PDF.',
                    'failures': failures}

        # ── Orden estable: curso ASC, nombre ASC, ID ASC, sesión 1S→2S ──
        ses_order = {'1S': 0, '2S': 1, 'M': 2}
        out_pdfs.sort(key=lambda x: (x[0], x[1], x[2], ses_order.get(x[3], 9)))

        # ── Mergear en un solo PDF ──
        try:
            writer = PdfWriter()
            for curso, nombre, sid, ses, pdf_path in out_pdfs:
                reader = PdfReader(str(pdf_path))
                for page in reader.pages:
                    writer.add_page(page)
            merged_buf = io.BytesIO()
            writer.write(merged_buf)
            writer.close()
            pdf_bytes = merged_buf.getvalue()
        except Exception as e:
            return {'success': False,
                    'error': f'Error al mergear PDFs: {e}',
                    'failures': failures}

        return {
            'success':      True,
            'zip_bytes':    pdf_bytes,   # nombre legacy: ahora es PDF
            'pdf_bytes':    pdf_bytes,
            'count':        len(out_pdfs),
            'failures':     len(failures),
            'roster_saved': roster_saved,
        }
